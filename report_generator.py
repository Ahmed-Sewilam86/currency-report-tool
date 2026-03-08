import pandas as pd
import logging
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def format_header(worksheet):
    """
    Apply formatting to header row.
    """
    header_fill = PatternFill(
        start_color="1E1E3A",
        end_color="1E1E3A",
        fill_type="solid"
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"


def auto_adjust_column_width(worksheet, dataframe):
    """
    Automatically adjust column width based on content.
    """
    for i, column in enumerate(dataframe.columns, 1):
        max_length = max(
            dataframe[column].astype(str).map(len).max(),
            len(column)
        )
        adjusted_width = max_length + 4
        worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width


def format_rate_column(worksheet, dataframe):
    """
    Apply number formatting to Exchange Rate column.
    """
    if "Exchange Rate" in dataframe.columns:
        col_index = dataframe.columns.get_loc("Exchange Rate") + 1
        column_letter = get_column_letter(col_index)

        for row in range(2, len(dataframe) + 2):
            worksheet[f"{column_letter}{row}"].number_format = "0.0000"
            worksheet[f"{column_letter}{row}"].alignment = Alignment(
                horizontal="center"
            )


def generate_excel_report(output_path: str, df: pd.DataFrame) -> None:
    """
    Generate formatted Excel report.

    Args:
        output_path (str): Path to save the report.
        df (pd.DataFrame): Processed exchange rates data.
    """
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Exchange Rates", index=False)

            worksheet = writer.book["Exchange Rates"]

            format_header(worksheet)
            auto_adjust_column_width(worksheet, df)
            format_rate_column(worksheet, df)

        logging.info("Excel report generated successfully.")

    except Exception as e:
        logging.error(f"Error generating report: {e}")
        raise